import sys
from pathlib import Path
root = Path(__file__).parent
sys.path.append(str(root.parent))
import pandas as pd
from openpyxl import Workbook
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from config import logger
from utils.data_tools import auto_fit_columns



def generate_excel_report(df: pd.DataFrame, stats_result: dict, report_path: str):
    logger.info("===== 开始生成Excel分析报告 =====")
    wb = Workbook()

    title_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=14)
    title_fill = PatternFill('solid', fgColor='003366')
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='0070C0')
    data_font = Font(name='微软雅黑', size=10)
    zebra_fill1 = PatternFill('solid', fgColor='EBF1F8')
    zebra_fill2 = PatternFill('solid', fgColor='FFFFFF')
    thin_gray = Side(style='thin', color='D9D9D9')
    h_border = Border(bottom=thin_gray)
    section_font = Font(name='微软雅黑', bold=True, size=12)
    section_fill = PatternFill('solid', fgColor='D9E1F2')

    # Sheet1 原始数据
    ws1 = wb.active
    ws1.title = '原始数据明细'
    ws1.merge_cells('A1:I1')
    ws1['A1'].value = '惠农网农产品行情原始数据'
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 40

    headers = ['品类', '产品名称', '产地', '省份', '原始价格文本', '价格(元/斤)', '涨跌幅(%)', '发布日期', '价格区间']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws1.row_dimensions[2].height = 30

    for row_idx, row_data in enumerate(df.itertuples(index=False), 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = h_border
            if col_idx in [1, 3, 4, 9]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [2, 5]:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = zebra_fill1 if (row_idx - 3) % 2 == 0 else zebra_fill2
    for row in range(3, len(df)+3):
        ws1.cell(row=row, column=6).number_format = '0.00'
        ws1.cell(row=row, column=7).number_format = '0.00'
        ws1.cell(row=row, column=8).number_format = 'YYYY-MM-DD'
    ws1.freeze_panes = 'A3'
    auto_fit_columns(ws1)

    # Sheet2 核心统计分析
    ws2 = wb.create_sheet('核心统计分析')
    current_row = 1
    ws2.merge_cells(f'A{current_row}:I{current_row}')
    ws2[f'A{current_row}'].value = '农产品价格核心统计分析报告'
    ws2[f'A{current_row}'].font = title_font
    ws2[f'A{current_row}'].fill = title_fill
    ws2[f'A{current_row}'].alignment = center_align
    ws2.row_dimensions[current_row].height = 40
    current_row += 2

    # 一、整体价格统计
    ws2.merge_cells(f'A{current_row}:B{current_row}')
    ws2[f'A{current_row}'].value = '一、整体价格统计'
    ws2[f'A{current_row}'].font = section_font
    ws2[f'A{current_row}'].fill = section_fill
    current_row += 1
    overall = stats_result['整体价格统计']
    for idx, (k, v) in enumerate(overall.items(), current_row):
        ws2.cell(row=idx, column=1, value=k).font = Font(name='微软雅黑', bold=True)
        ws2.cell(row=idx, column=2, value=v).font = data_font
    current_row = idx + 2

    # 二、各品类价格核心统计
    ws2.merge_cells(f'A{current_row}:I{current_row}')
    ws2[f'A{current_row}'].value = '二、各品类价格核心统计'
    ws2[f'A{current_row}'].font = section_font
    ws2[f'A{current_row}'].fill = section_fill
    current_row += 1
    cate_stats = stats_result['品类核心统计']
    cate_headers = ['品类', '样本数', '均价', '中位数', '标准差', '最低价', '最高价', '极差', '变异系数']
    for col, h in enumerate(cate_headers, 1):
        cell = ws2.cell(row=current_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    current_row += 1
    for cate, row_data in cate_stats.iterrows():
        ws2.cell(row=current_row, column=1, value=cate).font = Font(name='微软雅黑', bold=True)
        for col, v in enumerate(row_data, 2):
            ws2.cell(row=current_row, column=col, value=v).font = data_font
        current_row += 1
    current_row += 2

    # 三、产地覆盖统计
    ws2.merge_cells(f'A{current_row}:C{current_row}')
    ws2[f'A{current_row}'].value = '三、各品类产地覆盖情况'
    ws2[f'A{current_row}'].font = section_font
    ws2[f'A{current_row}'].fill = section_fill
    current_row += 1
    cover_headers = ['品类', '覆盖省份数', '产地总数']
    for col, h in enumerate(cover_headers, 1):
        cell = ws2.cell(row=current_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    current_row += 1
    cover_stats = stats_result['产地覆盖统计']
    for cate, row_data in cover_stats.iterrows():
        ws2.cell(row=current_row, column=1, value=cate).font = Font(name='微软雅黑', bold=True)
        for col, v in enumerate(row_data, 2):
            ws2.cell(row=current_row, column=col, value=v).font = data_font
        current_row += 1
    current_row += 2

    # 四、省份均价Top5
    ws2.merge_cells(f'A{current_row}:D{current_row}')
    ws2[f'A{current_row}'].value = '四、各品类省份均价 TOP5'
    ws2[f'A{current_row}'].font = section_font
    ws2[f'A{current_row}'].fill = section_fill
    current_row += 1
    prov_top5 = stats_result['省份均价Top5']
    for cate, prov_df in prov_top5.items():
        ws2.cell(row=current_row, column=1, value=f'【{cate}】').font = Font(name='微软雅黑', bold=True, color='0070C0')
        current_row += 1
        ws2.cell(row=current_row, column=1, value='排名').font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value='省份').font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        ws2.cell(row=current_row, column=3, value='均价(元/斤)').font = header_font
        ws2.cell(row=current_row, column=3).fill = header_fill
        ws2.cell(row=current_row, column=4, value='样本数').font = header_font
        ws2.cell(row=current_row, column=4).fill = header_fill
        for col in range(1,5):
            ws2.cell(row=current_row, column=col).alignment = center_align
        current_row += 1
        for rank, (prov, row_data) in enumerate(prov_df.iterrows(), 1):
            ws2.cell(row=current_row, column=1, value=rank).font = data_font
            ws2.cell(row=current_row, column=2, value=prov).font = data_font
            ws2.cell(row=current_row, column=3, value=row_data['均价']).font = data_font
            ws2.cell(row=current_row, column=4, value=row_data['样本数']).font = data_font
            for col in range(1,5):
                ws2.cell(row=current_row, column=col).alignment = center_align
            current_row += 1
        current_row += 1

    auto_fit_columns(ws2)

    # Sheet3 可视化图表
    ws3 = wb.create_sheet('可视化图表')
    ws3.merge_cells('A1:H1')
    ws3['A1'].value = '农产品价格数据可视化图表'
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 40

    def insert_image(ws, row, col, img_path, title, target_width=700):
        if not os.path.exists(img_path):
            logger.warning(f"图片 {img_path} 不存在，跳过")
            return row
        ws.cell(row=row, column=col, value=title).font = Font(name='微软雅黑', bold=True, size=12)
        img = XLImage(img_path)
        img_width = target_width
        img_height = int(img_width * (img.height / img.width))
        img.width = img_width
        img.height = img_height
        ws.add_image(img, f'{chr(64+col)}{row+1}')
        rows_taken = int(img_height / 15) + 3
        return row + rows_taken

    current_row = 3
    current_row = insert_image(ws3, current_row, 1, 'assets/output/各品类均价对比.png', '图1：各品类农产品均价对比')
    current_row = insert_image(ws3, current_row, 1, 'assets/output/价格区间分布.png', '图2：农产品价格区间整体分布')
    current_row = insert_image(ws3, current_row, 1, 'assets/output/品类价格波动箱线图.png', '图3：各品类农产品价格波动分布')
    current_row = insert_image(ws3, current_row, 1, 'assets/output/产品名称词云.png', '图4：产品名称词云图')
    current_row = insert_image(ws3, current_row, 1, 'assets/output/产品名称均价对比.png', '图5：Top20产品名称均价对比')
    current_row = insert_image(ws3, current_row, 1, 'assets/output/品类产品价格分布.png', '图6：各品类下产品名称价格分布')

    auto_fit_columns(ws3)
    wb.save(report_path)
    logger.info(f"Excel报告已保存至：{report_path}")
